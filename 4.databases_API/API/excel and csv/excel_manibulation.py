from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart,Reference

wb = load_workbook(filename='excel.xlsx')
ws = wb.active
for row in ws.iter_rows(min_row=1 , min_col=1, max_row=4, max_col=5, values_only=True):
    print(row)

'''data = [
    ['ahmed', 98, 97, 96, 96],
    ['ali', 93, 91, 89, 87],
    ['mohammed', 94, 93, 96, 91],
    ['nader', 90, 81, 82, 78]
]
wb = Workbook()

ws = wb.active

for row in data:
    ws.append(row)

for i in range(1, len(data)+1):
    ws.cell(row=i , column=6, value=f'=AVERAGE(B{i}:E{i})')

values = Reference(ws, min_col=1, min_row=1 , max_col=5 ,max_row=4)
names = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=4)

chart = BarChart()

chart.type = 'bar'
chart.title = 'Students Degree'
chart.x_axis.title = 'Student'
chart.y_axis.title = 'Degree'
chart.width = 20
chart.height = 15
chart.legend = None

chart.add_data(values)

ws.add_chart(chart, 'E15')

wb.save('excel.xlsx')'''